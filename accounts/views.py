"""Accounts views

Berisi view untuk registrasi, login, profile, manajemen user,
reset password via OTP, dan export laporan untuk admin.
"""

# Standard library
import json
import re
import random
import time
from datetime import datetime, timedelta
from calendar import monthrange

# Third-party
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (
    authenticate, login, logout, update_session_auth_hash, get_user_model
)
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils import timezone
from django.utils.timezone import now

# Local apps
from orders.models import Order, Promo
from services.models import Service
from .forms import CustomPasswordChangeForm, ProfileForm, CustomUserCreationForm
from .models import PasswordResetOTP
from .waha_service import WAHAHandler

User = get_user_model()


def register(request):
    """Registrasi akun baru dengan verifikasi WhatsApp"""
    
    print(f"=== REGISTER VIEW ===")
    print(f"Method: {request.method}")
    
    if request.method == 'POST':
        print("Processing POST request...")
        
        # Cek apakah ini step verifikasi OTP
        if request.POST.get('step') == 'verify_otp':
            print("Step: verify_otp")
            return verify_registration_otp(request)
        
        # Proses registrasi normal - KIRIM OTP
        print("Processing registration - sending OTP...")
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        address = request.POST.get('address', '').strip()
        phone = request.POST.get('phone', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        
        # Validasi
        errors = []
        if not first_name:
            errors.append("Nama depan wajib diisi!")
        if not username:
            errors.append("Username wajib diisi!")
        if not phone:
            errors.append("Nomor HP wajib diisi!")
        if password1 != password2:
            errors.append("Password tidak cocok!")
        if len(password1) < 8:
            errors.append("Password minimal 8 karakter!")
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect('accounts:register')
        
        # Cek username
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username sudah digunakan!")
            return redirect('accounts:register')
        
        # Format nomor HP
        phone_raw = ''.join(filter(str.isdigit, phone))
        if phone_raw.startswith('0'):
            phone_raw = '62' + phone_raw[1:]
        
        if not phone_raw.startswith('62') or len(phone_raw) < 10:
            messages.error(request, "Nomor HP tidak valid! Gunakan format 628xxxxxxxxxx")
            return redirect('accounts:register')
        
        # Cek nomor HP
        if User.objects.filter(phone=phone_raw).exists():
            messages.error(request, "Nomor HP sudah terdaftar!")
            return redirect('accounts:register')
        
        # =====================
        # GENERATE OTP
        # =====================
        otp_code = str(random.randint(100000, 999999))
        otp_expiry = time.time() + 300  # 5 menit
        
        # Simpan data registrasi ke session
        request.session['reg_data'] = {
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'address': address,
            'phone': phone_raw,
            'password': password1,
            'otp': otp_code,
            'otp_expiry': otp_expiry
        }
        
        # =====================
        # KIRIM OTP VIA WHATSAPP
        # =====================
        from .waha_service import WAHAHandler
        
        waha = WAHAHandler()
        message = f"""*Verifikasi Registrasi Menara Laundry*

Halo {first_name}!

Kode verifikasi Anda adalah:

*{otp_code}*

Kode ini berlaku selama 5 menit.

Jangan berikan kode ini kepada siapapun.

---
Menara Laundry - Solusi Laundry Praktis & Terpercaya"""
        
        try:
            success = waha.send_message(phone_raw, message)
            
            if success:
                messages.info(request, f"Kode OTP telah dikirim ke WhatsApp {phone_raw}")
                return render(request, 'accounts/verify_otp.html', {
                    'phone': phone_raw,
                    'step': 'verify'
                })
            else:
                messages.error(request, "Gagal mengirim kode verifikasi. Silakan coba lagi.")
                return redirect('accounts:register')
                
        except Exception as e:
            print(f"Error sending OTP: {e}")
            messages.error(request, "Terjadi kesalahan. Silakan coba lagi.")
            return redirect('accounts:register')
    
    # Handle GET request
    print("Showing registration form (GET request)")
    return render(request, 'accounts/register.html')

def clean_wa_id(wa_id):
    return re.sub(r"[^0-9]", "", wa_id)

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def link_whatsapp(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Gunakan POST"}, status=405)

    try:
        # =========================
        # PARSE REQUEST BODY
        # =========================
        data = {}

        if request.body:
            try:
                data = json.loads(request.body.decode('utf-8'))
            except Exception:
                data = {}
        else:
            data = request.POST.dict()

        # =========================
        # NORMALISASI PAYLOAD
        # =========================
        payload = data.get('payload', data)
        if not isinstance(payload, dict):
            return JsonResponse({
                "reply": "Format request tidak valid",
                "status": "invalid_payload"
            }, status=400)

        # =========================
        # AMBIL MESSAGE & WA ID
        # =========================
        message = (payload.get('body') or payload.get('message') or '').strip()
        message = re.sub(r"\s+", " ", message)  # normalize whitespace

        wa_id = payload.get('from') or payload.get('wa_id') or ''
        wa_id_clean = re.sub(r"[^0-9]", "", wa_id)

        # =========================
        # VALIDASI FORMAT COMMAND
        # =========================
        parts = message.split(" ", 1)

        if len(parts) != 2:
            return JsonResponse({
                "reply": "Format:\nLINK 628xxxxxxxx",
                "status": "invalid_format"
            }, status=400)

        command = parts[0].strip().upper()
        phone_raw = parts[1].strip()

        phone = re.sub(r"[^0-9]", "", phone_raw)

        # =========================
        # VALIDASI COMMAND
        # =========================
        if command != "LINK":
            return JsonResponse({
                "reply": "Perintah tidak dikenali",
                "status": "invalid_command"
            }, status=400)

        # =========================
        # NORMALISASI NOMOR
        # =========================
        if phone.startswith("0"):
            phone = "62" + phone[1:]

        # validasi ketat nomor Indonesia
        if not re.match(r"^62[0-9]{8,13}$", phone):
            return JsonResponse({
                "reply": "Nomor HP tidak valid. Gunakan format 628xxxxxxxx",
                "status": "invalid_phone"
            }, status=400)

        # =========================
        # CEK USER
        # =========================
        user = User.objects.filter(phone=phone).first()

        if not user:
            return JsonResponse({
                "reply": "Nomor tidak ditemukan",
                "status": "not_found"
            }, status=404)

        # =========================
        # CEK KONFLIK WA ID
        # =========================
        if wa_id_clean:
            conflict = User.objects.filter(wa_id=wa_id_clean).exclude(pk=user.pk).first()
            if conflict:
                return JsonResponse({
                    "reply": "WhatsApp sudah terhubung ke akun lain.",
                    "status": "conflict"
                }, status=409)

        # =========================
        # SIMPAN WA ID
        # =========================
        try:
            user.wa_id = wa_id  # SIMPAN RAW
            user.save()
        except Exception as e:
            return JsonResponse({
                "reply": "Gagal menyimpan data pengguna.",
                "status": "error",
                "detail": str(e)
            }, status=500)

        # =========================
        # SUCCESS RESPONSE
        # =========================
        return JsonResponse({
            "reply": "✅ WhatsApp berhasil terhubung",
            "status": "linked"
        }, status=200)

    except Exception as e:
        return JsonResponse({
            "reply": "Terjadi kesalahan internal",
            "status": "error",
            "detail": str(e)
        }, status=500)

def verify_registration_otp(request):
    """Verifikasi OTP untuk registrasi"""
    
    reg_data = request.session.get('reg_data')
    
    if not reg_data:
        messages.error(request, "Sesi registrasi tidak ditemukan. Silakan registrasi ulang.")
        return redirect('accounts:register')
    
    # Cek apakah OTP sudah kadaluarsa
    if time.time() > reg_data.get('otp_expiry', 0):
        messages.error(request, "Kode OTP sudah kadaluarsa. Silakan registrasi ulang.")
        # Hapus data session
        if 'reg_data' in request.session:
            del request.session['reg_data']
        return redirect('accounts:register')
    
    input_otp = request.POST.get('otp')
    
    if not input_otp:
        messages.error(request, "Silakan masukkan kode OTP.")
        return render(request, 'accounts/verify_otp.html', {
            'phone': reg_data.get('phone'),
            'step': 'verify'
        })
    
    if input_otp == reg_data.get('otp'):
        # =====================
        # BUAT AKUN BARU
        # =====================
        try:
            user = User.objects.create_user(
                username=reg_data['username'],
                password=reg_data['password'],
                first_name=reg_data['first_name'],
                last_name=reg_data['last_name'],
                address=reg_data['address'],
                phone=reg_data['phone'], 
            )
            
            # Simpan nomor HP
            user.phone = reg_data['phone']
            user.is_active = True
            user.save()
            
            # Hapus data session
            if 'reg_data' in request.session:
                del request.session['reg_data']
            
            messages.success(
                request, 
                "Akun berhasil dibuat! Silakan login dengan username dan password Anda."
            )
            return redirect('accounts:login')
            
        except Exception as e:
            print(f"Error membuat akun: {e}")
            messages.error(request, "Terjadi kesalahan saat membuat akun. Silakan coba lagi.")
            return redirect('accounts:register')
    else:
        messages.error(request, "Kode OTP salah! Silakan coba lagi.")
        return render(request, 'accounts/verify_otp.html', {
            'phone': reg_data.get('phone'),
            'step': 'verify'
        })


def user_login(request):
    """Login user dan arahkan berdasarkan role"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            messages.success(request, f"Login berhasil! Selamat datang, {user.first_name or user.username} 😊")

            if user.is_staff or user.is_superuser:
                # Admin
                return redirect('accounts:admin_dashboard')
            elif getattr(user, 'is_courier', False):
                # Kurir
                return redirect('courier:courier_dashboard')
            else:
                # Member biasa
                return redirect('accounts:home')
        else:
            messages.error(request, "Username atau password salah!")

    return render(request, 'accounts/login.html')


def home(request):
    from orders.models import Order
    from orders.models import Promo
    
    # Ambil promo
    promos = Promo.objects.filter(is_active=True).order_by('-created_at')
    
    # Hitung customer dari tabel User (yang sudah pasti ada datanya 4)
    total_customers = User.objects.count()
    
    # Hitung order dari tabel Order (yang sudah pasti ada datanya 3)
    total_orders = Order.objects.count()
    
    # DEBUG - cek di terminal
    print(f"CUSTOMERS: {total_customers}")  # Harusnya 4
    print(f"ORDERS: {total_orders}")        # Harusnya 3
    print(f"PROMOS: {promos.count()}")      # Harusnya 1
    
    return render(request, 'home.html', {
        'promos': promos,
        'total_customers': total_customers,
        'total_orders': total_orders,
    })

# ===============================
# 🔹 PROFILE VIEW
# ===============================
@login_required
def profile_view(request):
    """Halaman profil user dan ubah password"""
    user = request.user

    # Update profil
    if request.method == 'POST' and 'update_profile' in request.POST:
        profile_form = ProfileForm(request.POST, instance=user)
        if profile_form.is_valid():
            profile_form.save()
            messages.success(request, "Profil berhasil diperbarui!")
            return redirect('accounts:profile')
    else:
        profile_form = ProfileForm(instance=user)

    # Ganti password
    if request.method == 'POST' and 'update_password' in request.POST:
        password_form = CustomPasswordChangeForm(user=user, data=request.POST)
        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Password berhasil diganti!")
            return redirect('accounts:profile')
    else:
        password_form = PasswordChangeForm(user=user)

    context = {
        'profile_form': profile_form,
        'password_form': password_form,
    }
    return render(request, 'profile.html', context)

# accounts/views.py

@login_required
def admin_dashboard(request):
    """Halaman dashboard admin"""
    if not request.user.is_staff:
        return redirect('accounts:home')

    # 🔹 Data utama
    active_tab = request.GET.get('tab', 'dashboard')
    total_users = User.objects.count()
    total_orders = Order.objects.count()
    total_services = Service.objects.count()
    couriers = User.objects.filter(is_courier=True)
    total_couriers = couriers.count()

    # 🔹 Pendapatan (hanya yang sudah dibayar)
    paid_orders = Order.objects.filter(
        payment_status__in=["paid", "settlement"]
    )
    total_income = paid_orders.aggregate(total=Sum("price_total"))["total"] or 0
    today = now().date()
    today_income = paid_orders.filter(created_at__date=today).aggregate(total=Sum("price_total"))["total"] or 0
    total_transactions = paid_orders.count()

    # 🔹 Pendapatan 7 hari terakhir untuk grafik
    income_chart_labels = []
    income_chart_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        income_day = paid_orders.filter(created_at__date=day).aggregate(total=Sum("price_total"))["total"] or 0
        income_chart_labels.append(day.strftime("%d %b"))
        income_chart_data.append(float(income_day))

    # 🔹 Pesanan terbaru
    recent_orders_list = Order.objects.select_related(
        "customer", "service", "assigned_courier"
    ).order_by("-created_at")

    # 🔹 Transaksi terbaru
    recent_transactions_list = paid_orders.select_related("customer").order_by("-created_at")

    # Pagination
    orders_page_number = request.GET.get("orders_page", 1)
    transactions_page_number = request.GET.get("transactions_page", 1)

    orders_paginator = Paginator(recent_orders_list, 5)
    transactions_paginator = Paginator(recent_transactions_list, 5)

    recent_orders = orders_paginator.get_page(orders_page_number)
    recent_transactions = transactions_paginator.get_page(transactions_page_number)

    # 🔥 PERBAIKAN: Hitung pending_payments dengan benar
    pending_payments = Order.objects.filter(
        payment_status='waiting_confirmation'
    ).count()
    
    # Bukti pembayaran yang sudah dikonfirmasi
    confirmed_payments = Order.objects.filter(
        payment_status='paid'
    ).count()

    # 🔥 PERBAIKAN: recent_orders_with_proof - hanya yang ada proof
    recent_orders_with_proof = Order.objects.filter(
        payment_proof__isnull=False
    ).order_by('-created_at')[:10]

    context = {
        'active_tab': active_tab,
        "total_users": total_users,
        "total_orders": total_orders,
        "total_services": total_services,
        "total_couriers": total_couriers,
        "total_income": total_income,
        "today_income": today_income,
        "total_transactions": total_transactions,
        "recent_orders": recent_orders,
        'recent_orders_with_proof': recent_orders_with_proof,
        "recent_transactions": recent_transactions,
        "couriers": couriers,
        "income_chart_labels": income_chart_labels,
        "income_chart_data": income_chart_data,
        "orders_paginator": orders_paginator,
        "transactions_paginator": transactions_paginator,
        'pending_payments': pending_payments,  # 🔥 PASTIKAN INI ADA
        'confirmed_payments': confirmed_payments,
    }
    
    if active_tab == 'orders':
        context['active_tab'] = 'orders'

    return render(request, "accounts/admin_dashboard.html", context)

@login_required
def add_courier(request):
    if not request.user.is_staff:
        return redirect('accounts:home')

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        phone = request.POST.get("phone")
        
        # Validasi username
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username sudah digunakan.")
            return redirect('accounts:add_courier')
        
        # Validasi nomor HP
        if not phone:
            messages.error(request, "Nomor HP wajib diisi!")
            return redirect('accounts:add_courier')
        
        # Format nomor HP (hapus spasi, strip, dan karakter khusus)
        phone = ''.join(filter(str.isdigit, phone))
        
        # Jika dimulai dengan 0, ganti jadi 62
        if phone.startswith('0'):
            phone = '62' + phone[1:]
        
        # Validasi nomor HP hanya angka dan panjangnya sesuai
        if not phone.isdigit():
            messages.error(request, "Nomor HP harus berupa angka!")
            return redirect('accounts:add_courier')
        
        if len(phone) < 10 or len(phone) > 15:
            messages.error(request, "Nomor HP harus antara 10-15 digit!")
            return redirect('accounts:add_courier')
        
        if not phone.startswith('62'):
            messages.error(request, "Nomor HP harus dimulai dengan 62 (contoh: 628123456789)")
            return redirect('accounts:add_courier')
        
        # Cek nomor HP sudah terdaftar
        if User.objects.filter(phone=phone).exists():
            messages.error(request, "Nomor HP sudah digunakan!")
            return redirect('accounts:add_courier')
        
        # Buat akun kurir
        try:
            courier = User.objects.create_user(
                username=username,
                password=password,
                phone=phone,
                is_courier=True,
                is_customer=False,
                is_active=True  # Langsung aktif
            )
            
            messages.success(request, f"Kurir '{courier.username}' berhasil ditambahkan!")
            
            # Opsional: Kirim notifikasi WhatsApp ke kurir
            try:
                from .waha_service import WAHAHandler
                waha = WAHAHandler()
                message = f"""🎉 *Selamat! Anda Telah Menjadi Kurir*

Halo {username}!

Anda telah ditambahkan sebagai kurir di Menara Laundry.

━━━━━━━━━━━━━━━━━━━━
📋 *Informasi Akun*
━━━━━━━━━━━━━━━━━━━━

👤 *Username:* {username}
🔑 *Password:* {password}
📱 *Nomor HP:* {phone}

━━━━━━━━━━━━━━━━━━━━
🔗 *Link Login:*
━━━━━━━━━━━━━━━━━━━━

https://www.menaralaundry.site/accounts/login/

━━━━━━━━━━━━━━━━━━━━

Segera ganti password setelah login untuk keamanan.

---
Menara Laundry"""
                
                waha.send_message(phone, message)
            except Exception as e:
                print(f"Gagal kirim notifikasi: {e}")
                
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('accounts:add_courier')
        
        return redirect('accounts:manage_users')

    return render(request, 'accounts/add_courier.html')

def admin_required(user):
    return user.is_staff

@login_required
@user_passes_test(admin_required)
def manage_users(request):
    users = User.objects.all()

    # Pisahkan Karyawan (Admin & Kurir) dan Member
    staff_users_list = [u for u in users if u.is_staff or getattr(u, "is_courier", False)]
    member_users_list = [u for u in users if not u.is_staff and not getattr(u, "is_courier", False)]

    # Pagination: 10 user per halaman
    staff_paginator = Paginator(staff_users_list, 10)
    member_paginator = Paginator(member_users_list, 10)

    staff_page_number = request.GET.get('staff_page', 1)
    member_page_number = request.GET.get('member_page', 1)

    staff_users = staff_paginator.get_page(staff_page_number)
    member_users = member_paginator.get_page(member_page_number)

    return render(request, 'accounts/manage_users.html', {
        'staff_users': staff_users,
        'member_users': member_users,
    })


@login_required
@user_passes_test(admin_required)
def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        
        # Ambil data dari form untuk validasi awal
        username = request.POST.get('username')
        phone = request.POST.get('phone')
        
        # =====================
        # VALIDASI SEBELUM FORM VALIDATION
        # =====================
        
        # Validasi username
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username sudah digunakan.")
            return redirect('accounts:add_user')
        
        # Validasi nomor HP
        if not phone:
            messages.error(request, "Nomor HP wajib diisi!")
            return redirect('accounts:add_user')
        
        # Format nomor HP (hapus spasi, strip, dan karakter khusus)
        phone_raw = ''.join(filter(str.isdigit, phone))
        
        # Jika dimulai dengan 0, ganti jadi 62
        if phone_raw.startswith('0'):
            phone_raw = '62' + phone_raw[1:]
        
        # Validasi nomor HP hanya angka dan panjangnya sesuai
        if not phone_raw.isdigit():
            messages.error(request, "Nomor HP harus berupa angka!")
            return redirect('accounts:add_user')
        
        if len(phone_raw) < 10 or len(phone_raw) > 15:
            messages.error(request, "Nomor HP harus antara 10-15 digit!")
            return redirect('accounts:add_user')
        
        if not phone_raw.startswith('62'):
            messages.error(request, "Nomor HP harus dimulai dengan 62 (contoh: 628123456789)")
            return redirect('accounts:add_user')
        
        # Cek nomor HP sudah terdaftar
        if User.objects.filter(phone=phone_raw).exists():
            messages.error(request, "Nomor HP sudah digunakan!")
            return redirect('accounts:add_user')
        
        # =====================
        # PROSES FORM
        # =====================
        
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = True  # langsung aktif
            user.phone = phone_raw  # gunakan nomor yang sudah diformat
            user.save()
            
            # =====================
            # KIRIM NOTIFIKASI VIA WHATSAPP
            # =====================
            from .waha_service import WAHAHandler
            
            waha = WAHAHandler()
            
            # Format pesan WhatsApp (bukan OTP)
            message = f"""🎉 *Selamat! Akun Anda Telah Dibuat*

Halo {user.first_name or user.username}!

Akun Menara Laundry Anda telah berhasil dibuat oleh Admin.

━━━━━━━━━━━━━━━━━━━━
📋 *Informasi Akun*
━━━━━━━━━━━━━━━━━━━━

👤 *Username:* {user.username}
🔑 *Password:* (Password yang Anda daftarkan)
📱 *Nomor HP:* {phone_raw}

━━━━━━━━━━━━━━━━━━━━
🔗 *Link Login:*
━━━━━━━━━━━━━━━━━━━━

https://www.menaralaundry.site/accounts/login/

━━━━━━━━━━━━━━━━━━━━
💡 *Tips:*
━━━━━━━━━━━━━━━━━━━━

1. Simpan username dan password Anda dengan aman
2. Jangan berikan informasi akun kepada siapapun
3. Segera ganti password setelah login untuk keamanan

Jika ada kendala, silakan hubungi admin.

---
Menara Laundry - Solusi Laundry Praktis & Terpercaya"""
            
            # Kirim notifikasi WhatsApp
            try:
                success = waha.send_message(phone_raw, message)
                
                if success:
                    messages.success(
                        request, 
                        f"User '{user.username}' berhasil ditambahkan. Notifikasi telah dikirim ke WhatsApp {phone_raw}."
                    )
                else:
                    messages.warning(
                        request, 
                        f"User '{user.username}' berhasil ditambahkan, tapi gagal mengirim notifikasi WhatsApp."
                    )
            except Exception as e:
                print(f"Error kirim WA: {e}")
                messages.warning(
                    request, 
                    f"User '{user.username}' berhasil ditambahkan, tapi notifikasi WhatsApp gagal dikirim."
                )
            
            return redirect('accounts:manage_users')
        else:
            # Jika form tidak valid, tampilkan error dari form
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return redirect('accounts:add_user')
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'accounts/add_user.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, f"Pengguna {user.username} berhasil dihapus.")
    return redirect('accounts:manage_users')

def password_reset_otp(request):
    """
    Reset password menggunakan OTP via WhatsApp
    """
    
    # GET request - Tampilkan form input nomor HP
    if request.method == "GET":
        return render(request, "accounts/password_reset_otp.html", {"step": "email"})
    
    # POST request - Proses berdasarkan step
    if request.method == "POST":
        step = request.POST.get("step")
        
        print(f"Password reset step: {step}")  # Debugging
        
        # =====================
        # STEP 1: KIRIM OTP VIA WHATSAPP
        # =====================
        if step == "send_otp":
            phone = request.POST.get("phone", "").strip()
            
            if not phone:
                messages.error(request, "Nomor HP wajib diisi!")
                return render(request, "accounts/password_reset_otp.html", {"step": "email"})
            
            # Format nomor HP
            phone = ''.join(filter(str.isdigit, phone))
            if phone.startswith('0'):
                phone = '62' + phone[1:]
            
            # Cari user berdasarkan nomor HP
            user = User.objects.filter(phone=phone).first()
            
            if user:
                # Hapus OTP lama yang belum digunakan
                PasswordResetOTP.objects.filter(user=user, is_used=False).delete()
                
                # Generate OTP 6 digit
                otp = str(random.randint(100000, 999999))
                PasswordResetOTP.objects.create(user=user, otp=otp)
                
                # Kirim via WAHA
                try:
                    waha = WAHAHandler()
                    
                    message = f"""🔐 *Reset Password Menara Laundry*

Halo {user.first_name or user.username}!

Kami menerima permintaan reset password.

━━━━━━━━━━━━━━━━━━━━
🔑 *Kode OTP Anda:*
*{otp}*
━━━━━━━━━━━━━━━━━━━━

⏰ Kode ini berlaku *5 menit*

Jika Anda tidak meminta reset password, abaikan pesan ini.

---
Menara Laundry"""
                    
                    success = waha.send_message(phone, message)
                    
                    if success:
                        # Simpan phone di session untuk keperluan verifikasi
                        request.session['reset_phone'] = phone
                        return render(
                            request,
                            "accounts/password_reset_otp.html",
                            {"step": "verify", "phone": phone}
                        )
                    else:
                        messages.error(
                            request,
                            "❌ Gagal mengirim OTP. Silakan coba lagi."
                        )
                        return render(request, "accounts/password_reset_otp.html", {"step": "email"})
                        
                except Exception as e:
                    print(f"Error sending WA: {e}")
                    messages.error(request, "❌ Gagal mengirim OTP. Silakan coba lagi.")
                    return render(request, "accounts/password_reset_otp.html", {"step": "email"})
            else:
                # Untuk keamanan, tetap tampilkan pesan sukses
                messages.info(
                    request,
                    "ℹ️ Jika nomor HP terdaftar, OTP akan dikirim dalam beberapa saat."
                )
                return render(request, "accounts/password_reset_otp.html", {"step": "email"})
        
        # =====================
        # STEP 2: VERIFIKASI OTP DAN RESET PASSWORD
        # =====================
        elif step == "verify_otp":
            otp = request.POST.get("otp", "").strip()
            password = request.POST.get("password", "")
            confirm_password = request.POST.get("confirm_password", "")
            phone = request.session.get('reset_phone', '')
            
            print(f"Verifying OTP: {otp} for phone: {phone}")  # Debugging
            
            # Validasi OTP
            if not otp:
                messages.error(request, "Kode OTP wajib diisi!")
                return render(
                    request,
                    "accounts/password_reset_otp.html",
                    {"step": "verify", "phone": phone}
                )
            
            # Validasi password
            if not password:
                messages.error(request, "Password baru wajib diisi!")
                return render(
                    request,
                    "accounts/password_reset_otp.html",
                    {"step": "verify", "phone": phone}
                )
            
            if password != confirm_password:
                messages.error(request, "Password dan konfirmasi tidak cocok!")
                return render(
                    request,
                    "accounts/password_reset_otp.html",
                    {"step": "verify", "phone": phone}
                )
            
            if len(password) < 8:
                messages.error(request, "Password minimal 8 karakter!")
                return render(
                    request,
                    "accounts/password_reset_otp.html",
                    {"step": "verify", "phone": phone}
                )
            
            # Cek OTP di database (belum digunakan dan belum expired)
            try:
                # Cari OTP yang valid
                record = PasswordResetOTP.objects.filter(
                    otp=otp,
                    is_used=False
                ).first()
                
                if not record:
                    messages.error(request, "❌ Kode OTP tidak valid!")
                    return render(
                        request,
                        "accounts/password_reset_otp.html",
                        {"step": "verify", "phone": phone}
                    )
                
                # Cek apakah OTP sudah expired
                if record.is_expired():
                    messages.error(request, "❌ Kode OTP sudah kadaluarsa! Silakan minta OTP baru.")
                    record.delete()  # Hapus OTP yang expired
                    return render(request, "accounts/password_reset_otp.html", {"step": "email"})
                
                # Reset password
                user = record.user
                user.set_password(password)
                user.save()
                
                # Tandai OTP sebagai sudah digunakan
                record.use_otp()
                
                # Hapus session
                if 'reset_phone' in request.session:
                    del request.session['reset_phone']
                
                messages.success(
                    request,
                    "✅ Password berhasil direset! Silakan login dengan password baru."
                )
                return redirect("accounts:login")
                
            except Exception as e:
                print(f"Error verifying OTP: {e}")
                messages.error(request, "❌ Terjadi kesalahan. Silakan coba lagi.")
                return render(
                    request,
                    "accounts/password_reset_otp.html",
                    {"step": "verify", "phone": phone}
                )
        
        else:
            # Step tidak dikenal
            messages.error(request, "Terjadi kesalahan. Silakan coba lagi.")
            return render(request, "accounts/password_reset_otp.html", {"step": "email"})
    
    # Fallback
    return render(request, "accounts/password_reset_otp.html", {"step": "email"})

# accounts/views.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from django.utils import timezone
from datetime import datetime
from calendar import monthrange
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
import io
import requests
from PIL import Image as PILImage


def resize_image_professional(image_data, max_width=600, max_height=400, quality=92):
    """
    Resize gambar dengan kualitas profesional untuk Excel
    """
    try:
        if isinstance(image_data, bytes):
            img = PILImage.open(io.BytesIO(image_data))
        else:
            img = PILImage.open(image_data)
        
        orig_width, orig_height = img.size
        ratio = orig_width / orig_height
        
        if ratio > 1:
            new_width = min(max_width, orig_width)
            new_height = new_width / ratio
            if new_height > max_height:
                new_height = max_height
                new_width = new_height * ratio
        else:
            new_height = min(max_height, orig_height)
            new_width = new_height * ratio
            if new_width > max_width:
                new_width = max_width
                new_height = new_width / ratio
        
        new_width = int(new_width)
        new_height = int(new_height)
        
        img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
        
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        output.seek(0)
        
        return output, new_width, new_height
        
    except Exception as e:
        print(f"Resize error: {e}")
        return None, 0, 0

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_orders_excel(request, year=None, month=None):
    """
    Export data pesanan ke Excel untuk rekap bulanan (hanya status pembayaran PAID/LUNAS)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as ExcelImage
    from calendar import monthrange
    import requests
    from io import BytesIO
    from PIL import Image
    
    now = timezone.now()
    target_year = year if year else now.year
    target_month = month if month else now.month
    
    first_day = datetime(target_year, target_month, 1)
    last_day = datetime(target_year, target_month, monthrange(target_year, target_month)[1])
    
    orders = Order.objects.filter(
        created_at__date__gte=first_day,
        created_at__date__lte=last_day,
        payment_status='paid'
    ).order_by('-created_at').select_related('customer', 'service', 'assigned_courier')
    
    bulan_names = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
        5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
        9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }
    
    wb = Workbook()
    
    # ========== STYLE DEFINITIONS ==========
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    subtitle_font = Font(name='Arial', size=10, italic=True, color='6B7280')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    
    # Fungsi untuk mendapatkan order display
    def get_order_display(order):
        return order.order_number if order.order_number else f"#{order.id}"
    
    # ============================================================
    # SHEET 1: REKAP BULANAN (MAIN)
    # ============================================================
    ws_rekap = wb.active
    ws_rekap.title = f"Rekap {bulan_names[target_month]} {target_year}"
    
    try:
        ws_rekap.page_setup.orientation = 'landscape'
        ws_rekap.page_setup.paperSize = '9'
        ws_rekap.page_setup.fitToPage = True
        ws_rekap.page_setup.fitToWidth = 1
        ws_rekap.page_setup.fitToHeight = 0
        ws_rekap.page_setup.horizontalCentered = True
        ws_rekap.page_setup.verticalCentered = True
        ws_rekap.print_area = f'A1:N{len(orders) + 15}'
        ws_rekap.print_title_rows = '1:10'
    except:
        pass
    
    # --- HEADER ---
    ws_rekap.merge_cells('A1:N1')
    ws_rekap['A1'] = f"MENARA LAUNDRY"
    ws_rekap['A1'].font = Font(name='Arial', size=16, bold=True, color='1E40AF')
    ws_rekap['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_rekap.row_dimensions[1].height = 30
    
    ws_rekap.merge_cells('A2:N2')
    ws_rekap['A2'] = f"LAPORAN REKAP TRANSAKSI BULANAN"
    ws_rekap['A2'].font = Font(name='Arial', size=12, bold=True)
    ws_rekap['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_rekap.merge_cells('A3:N3')
    ws_rekap['A3'] = f"{bulan_names[target_month].upper()} {target_year}"
    ws_rekap['A3'].font = Font(name='Arial', size=11, bold=True, color='4B5563')
    ws_rekap['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_rekap.merge_cells('A4:N4')
    ws_rekap['A4'] = f"Periode: {first_day.strftime('%d %B %Y')} - {last_day.strftime('%d %B %Y')}"
    ws_rekap['A4'].font = subtitle_font
    ws_rekap['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_rekap.merge_cells('A5:N5')
    ws_rekap['A5'] = "✅ STATUS PEMBAYARAN: LUNAS (PAID)"
    ws_rekap['A5'].font = Font(name='Arial', size=10, bold=True, color='065f46')
    ws_rekap['A5'].fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    ws_rekap['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws_rekap.row_dimensions[5].height = 22
    
    # --- SUMMARY ---
    total_orders = len(orders)
    total_keseluruhan = sum(float(o.price_total) for o in orders)
    total_ongkir = sum(float(o.shipping_cost or 0) for o in orders)
    total_diskon = sum(float(o.discount_amount or 0) for o in orders)
    total_berat = sum(float(o.weight or 0) for o in orders)
    
    stats_row = 7
    ws_rekap.merge_cells(f'A{stats_row}:G{stats_row}')
    ws_rekap[f'A{stats_row}'] = f"📊 TOTAL PESANAN: {total_orders}"
    ws_rekap[f'A{stats_row}'].font = Font(bold=True, size=10)
    ws_rekap[f'A{stats_row}'].fill = blue_fill
    
    ws_rekap.merge_cells(f'H{stats_row}:N{stats_row}')
    ws_rekap[f'H{stats_row}'] = f"💰 TOTAL PENDAPATAN: Rp {total_keseluruhan:,.0f}".replace(',', '.')
    ws_rekap[f'H{stats_row}'].font = Font(bold=True, size=10, color='065f46')
    ws_rekap[f'H{stats_row}'].fill = blue_fill
    
    stats_row += 1
    ws_rekap.merge_cells(f'A{stats_row}:G{stats_row}')
    ws_rekap[f'A{stats_row}'] = f"⚖️ TOTAL BERAT: {total_berat:,.1f} kg".replace(',', '.')
    ws_rekap[f'A{stats_row}'].font = Font(bold=True, size=10)
    ws_rekap[f'A{stats_row}'].fill = blue_fill
    
    ws_rekap.merge_cells(f'H{stats_row}:N{stats_row}')
    ws_rekap[f'H{stats_row}'] = f"📦 RATA-RATA: Rp {(total_keseluruhan/total_orders if total_orders > 0 else 0):,.0f}".replace(',', '.')
    ws_rekap[f'H{stats_row}'].font = Font(bold=True, size=10)
    ws_rekap[f'H{stats_row}'].fill = blue_fill
    
    # --- TABLE HEADER ---
    # PERBAIKAN: Ganti 'ID ORDER' menjadi 'NO. ORDER'
    headers = ['NO', 'NO. ORDER', 'TANGGAL', 'PELANGGAN', 'TELEPON', 'LAYANAN', 
               'BERAT', 'SUB TOTAL', 'ONGKIR', 'DISKON', 'TOTAL', 'STATUS', 'KURIR', 'BUKTI']
    columns_width = [5, 18, 15, 18, 13, 20, 10, 14, 12, 12, 14, 12, 15, 12]
    
    header_row = 10
    for col, (header, width) in enumerate(zip(headers, columns_width), start=1):
        cell = ws_rekap.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border
        ws_rekap.column_dimensions[get_column_letter(col)].width = width
        ws_rekap.row_dimensions[header_row].height = 25
    
    # --- DATA ---
    for idx, order in enumerate(orders, start=1):
        row = header_row + idx
        
        subtotal = float(order.price_total) + float(order.discount_amount or 0) - float(order.shipping_cost or 0)
        row_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid') if idx % 2 == 0 else None
        
        ws_rekap.cell(row=row, column=1, value=idx).border = thin_border
        ws_rekap.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        if row_fill: ws_rekap.cell(row=row, column=1).fill = row_fill
        
        # PERBAIKAN: Gunakan order_number
        order_display = get_order_display(order)
        ws_rekap.cell(row=row, column=2, value=order_display).border = thin_border
        ws_rekap.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        if row_fill: ws_rekap.cell(row=row, column=2).fill = row_fill
        
        ws_rekap.cell(row=row, column=3, value=order.created_at.strftime('%d/%m/%Y')).border = thin_border
        ws_rekap.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        if row_fill: ws_rekap.cell(row=row, column=3).fill = row_fill
        
        ws_rekap.cell(row=row, column=4, value=order.customer.get_full_name() or order.customer.username).border = thin_border
        if row_fill: ws_rekap.cell(row=row, column=4).fill = row_fill
        
        ws_rekap.cell(row=row, column=5, value=order.customer.phone or '-').border = thin_border
        ws_rekap.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        if row_fill: ws_rekap.cell(row=row, column=5).fill = row_fill
        
        ws_rekap.cell(row=row, column=6, value=order.service.name if order.service else '-').border = thin_border
        if row_fill: ws_rekap.cell(row=row, column=6).fill = row_fill
        
        ws_rekap.cell(row=row, column=7, value=float(order.weight or 0)).border = thin_border
        ws_rekap.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        ws_rekap.cell(row=row, column=7).number_format = '#,##0.0'
        if row_fill: ws_rekap.cell(row=row, column=7).fill = row_fill
        
        ws_rekap.cell(row=row, column=8, value=round(subtotal, 2)).border = thin_border
        ws_rekap.cell(row=row, column=8).number_format = '#,##0'
        if row_fill: ws_rekap.cell(row=row, column=8).fill = row_fill
        
        ws_rekap.cell(row=row, column=9, value=float(order.shipping_cost or 0)).border = thin_border
        ws_rekap.cell(row=row, column=9).number_format = '#,##0'
        if row_fill: ws_rekap.cell(row=row, column=9).fill = row_fill
        
        ws_rekap.cell(row=row, column=10, value=float(order.discount_amount or 0)).border = thin_border
        ws_rekap.cell(row=row, column=10).number_format = '#,##0'
        if row_fill: ws_rekap.cell(row=row, column=10).fill = row_fill
        
        ws_rekap.cell(row=row, column=11, value=float(order.price_total)).border = thin_border
        ws_rekap.cell(row=row, column=11).font = Font(bold=True)
        ws_rekap.cell(row=row, column=11).number_format = '#,##0'
        if row_fill: ws_rekap.cell(row=row, column=11).fill = row_fill
        
        status_cell = ws_rekap.cell(row=row, column=12, value=order.get_order_status_display())
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal='center')
        if order.order_status == 'delivered':
            status_cell.fill = green_fill
            status_cell.font = Font(color='065f46', bold=True)
        elif order.order_status == 'cancelled':
            status_cell.fill = red_fill
            status_cell.font = Font(color='991B1B', bold=True)
        elif row_fill:
            status_cell.fill = row_fill
        
        ws_rekap.cell(row=row, column=13, value=order.assigned_courier.username if order.assigned_courier else '-').border = thin_border
        ws_rekap.cell(row=row, column=13).alignment = Alignment(horizontal='center')
        if row_fill: ws_rekap.cell(row=row, column=13).fill = row_fill
        
        proof_cell = ws_rekap.cell(row=row, column=14)
        proof_cell.border = thin_border
        proof_cell.alignment = Alignment(horizontal='center')
        if order.payment_proof:
            proof_cell.value = "✅ Ada"
            proof_cell.font = Font(color='065f46', bold=True)
            proof_cell.fill = green_fill
        else:
            proof_cell.value = "❌ Tidak"
            proof_cell.font = Font(color='991B1B')
            proof_cell.fill = red_fill
        if row_fill and not order.payment_proof:
            proof_cell.fill = red_fill
    
    # --- FOOTER ---
    footer_row = header_row + total_orders + 2
    
    ws_rekap.merge_cells(f'A{footer_row}:F{footer_row}')
    ws_rekap[f'A{footer_row}'] = "TOTAL KESELURUHAN"
    ws_rekap[f'A{footer_row}'].font = Font(bold=True, size=11)
    ws_rekap[f'A{footer_row}'].alignment = Alignment(horizontal='right')
    ws_rekap[f'A{footer_row}'].border = thick_border
    ws_rekap[f'A{footer_row}'].fill = yellow_fill
    
    cell = ws_rekap.cell(row=footer_row, column=7, value=round(total_berat, 2))
    cell.font = Font(bold=True)
    cell.border = thick_border
    cell.number_format = '#,##0.0'
    cell.fill = yellow_fill
    
    cell = ws_rekap.cell(row=footer_row, column=8, value=round(total_keseluruhan + total_diskon - total_ongkir, 2))
    cell.font = Font(bold=True)
    cell.border = thick_border
    cell.number_format = '#,##0'
    cell.fill = yellow_fill
    
    cell = ws_rekap.cell(row=footer_row, column=9, value=round(total_ongkir, 2))
    cell.font = Font(bold=True)
    cell.border = thick_border
    cell.number_format = '#,##0'
    cell.fill = yellow_fill
    
    cell = ws_rekap.cell(row=footer_row, column=10, value=round(total_diskon, 2))
    cell.font = Font(bold=True)
    cell.border = thick_border
    cell.number_format = '#,##0'
    cell.fill = yellow_fill
    
    cell = ws_rekap.cell(row=footer_row, column=11, value=round(total_keseluruhan, 2))
    cell.font = Font(bold=True, size=11, color='1E40AF')
    cell.border = thick_border
    cell.number_format = '#,##0'
    cell.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    
    info_row = footer_row + 2
    ws_rekap.merge_cells(f'A{info_row}:N{info_row}')
    ws_rekap[f'A{info_row}'] = f"Dicetak: {timezone.now().strftime('%d %B %Y %H:%M:%S')} | User: {request.user.username} | Jumlah: {total_orders} transaksi"
    ws_rekap[f'A{info_row}'].font = Font(size=8, color='6B7280', italic=True)
    ws_rekap[f'A{info_row}'].alignment = Alignment(horizontal='center')
    
    # ============================================================
    # SHEET 2: BUKTI PEMBAYARAN + GALLERY (GABUNGAN)
    # ============================================================
    ws_proof = wb.create_sheet("Bukti & Gallery")
    
    try:
        ws_proof.page_setup.orientation = 'portrait'
        ws_proof.page_setup.paperSize = '9'
        ws_proof.page_setup.fitToPage = True
        ws_proof.page_setup.fitToWidth = 1
    except:
        pass
    
    # ===== HEADER =====
    ws_proof.merge_cells('A1:G1')
    ws_proof['A1'] = "📸 BUKTI & GALLERY PEMBAYARAN"
    ws_proof['A1'].font = Font(size=14, bold=True, color='1E40AF')
    ws_proof['A1'].alignment = Alignment(horizontal='center')
    
    ws_proof.merge_cells('A2:G2')
    ws_proof['A2'] = f"{bulan_names[target_month]} {target_year}"
    ws_proof['A2'].font = Font(size=10, italic=True)
    ws_proof['A2'].alignment = Alignment(horizontal='center')
    
    ws_proof.merge_cells('A3:G3')
    ws_proof['A3'] = "Status: LUNAS (PAID) | Klik gambar untuk melihat lebih besar"
    ws_proof['A3'].font = Font(size=10, color='065f46')
    ws_proof['A3'].fill = green_fill
    ws_proof['A3'].alignment = Alignment(horizontal='center')
    
    # ===== HEADER TABLE =====
    # PERBAIKAN: Ganti 'ID ORDER' menjadi 'NO. ORDER'
    proof_headers = ['NO', 'NO. ORDER', 'PELANGGAN', 'TANGGAL BAYAR', 'TOTAL', 'BUKTI', 'STATUS']
    proof_widths = [5, 20, 25, 17, 15, 35, 15]
    
    for col, (header, width) in enumerate(zip(proof_headers, proof_widths), start=1):
        cell = ws_proof.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thick_border
        ws_proof.column_dimensions[get_column_letter(col)].width = width
    
    # ===== DATA =====
    proof_row = 6
    proof_count = 0
    
    for idx, order in enumerate(orders, start=1):
        proof_count += 1
        
        # Set row height untuk gambar
        if order.payment_proof:
            ws_proof.row_dimensions[proof_row].height = 150
        else:
            ws_proof.row_dimensions[proof_row].height = 30
        
        # NO
        cell = ws_proof.cell(row=proof_row, column=1, value=idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # PERBAIKAN: NO. ORDER menggunakan order_number
        order_display = get_order_display(order)
        cell = ws_proof.cell(row=proof_row, column=2, value=order_display)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        
        # PELANGGAN
        cell = ws_proof.cell(row=proof_row, column=3, value=order.customer.get_full_name() or order.customer.username)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # TANGGAL BAYAR
        cell = ws_proof.cell(row=proof_row, column=4, value=order.payment_date.strftime('%d/%m/%Y %H:%M') if order.payment_date else '-')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # TOTAL
        cell = ws_proof.cell(row=proof_row, column=5, value=float(order.price_total))
        cell.border = thin_border
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(bold=True)
        
        # ===== BUKTI (GAMBAR) - RATA TENGAH =====
        img_cell = ws_proof.cell(row=proof_row, column=6)
        img_cell.border = thin_border
        img_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if order.payment_proof and order.payment_proof.url:
            try:
                if order.payment_proof.url.startswith('http'):
                    image_url = order.payment_proof.url
                else:
                    image_url = request.build_absolute_uri(order.payment_proof.url)
                
                response = requests.get(image_url, timeout=15)
                
                if response.status_code == 200:
                    # Resize dengan proporsi yang benar
                    img_bytes, new_width, new_height = resize_image_professional(
                        response.content,
                        max_width=500,
                        max_height=300,
                        quality=92
                    )
                    
                    if img_bytes:
                        img_excel = ExcelImage(img_bytes)
                        
                        # Hitung ukuran untuk Excel
                        excel_width = int(new_width * 0.8)
                        excel_height = int(new_height * 0.8)
                        
                        # Batasi ukuran maksimal
                        max_excel_width = 350
                        max_excel_height = 250
                        
                        if excel_width > max_excel_width:
                            ratio = max_excel_width / excel_width
                            excel_width = max_excel_width
                            excel_height = int(excel_height * ratio)
                        
                        if excel_height > max_excel_height:
                            ratio = max_excel_height / excel_height
                            excel_height = max_excel_height
                            excel_width = int(excel_width * ratio)
                        
                        # Pastikan tidak terlalu kecil
                        if excel_width < 150:
                            excel_width = 150
                            excel_height = int(excel_width * (new_height / new_width))
                        
                        img_excel.width = excel_width
                        img_excel.height = excel_height
                        
                        # Tambahkan gambar ke cell
                        cell_coord = f'F{proof_row}'
                        ws_proof.add_image(img_excel, cell_coord)
                        
                        # Kosongkan value cell
                        img_cell.value = ""
                    else:
                        img_cell.value = "❌ Gagal resize"
                        img_cell.font = Font(color='991B1B', size=9)
                        img_cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    img_cell.value = "❌ Gagal load"
                    img_cell.font = Font(color='991B1B', size=9)
                    img_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
            except Exception as e:
                print(f"Error: {e}")
                img_cell.value = "⚠️ Error"
                img_cell.font = Font(color='F59E0B', size=9)
                img_cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            img_cell.value = "Tidak ada bukti"
            img_cell.font = Font(color='9CA3AF', size=9, italic=True)
            img_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ===== STATUS =====
        status_cell = ws_proof.cell(row=proof_row, column=7)
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        if order.payment_proof:
            status_cell.value = "✅ Ada"
            status_cell.font = Font(color='065f46', bold=True, size=10)
            status_cell.fill = green_fill
        else:
            status_cell.value = "❌ Tidak"
            status_cell.font = Font(color='991B1B', size=10)
            status_cell.fill = red_fill
        
        proof_row += 1
    
    # ============================================================
    # SHEET 3: REKAP PER LAYANAN
    # ============================================================
    ws_service = wb.create_sheet("Rekap per Layanan")
    
    try:
        ws_service.page_setup.orientation = 'portrait'
        ws_service.page_setup.paperSize = '9'
        ws_service.page_setup.fitToPage = True
        ws_service.page_setup.fitToWidth = 1
    except:
        pass
    
    ws_service.merge_cells('A1:E1')
    ws_service['A1'] = f"REKAP PER LAYANAN - {bulan_names[target_month]} {target_year}"
    ws_service['A1'].font = Font(size=12, bold=True, color='1E40AF')
    ws_service['A1'].alignment = Alignment(horizontal='center')
    
    service_headers = ['No', 'Nama Layanan', 'Jenis', 'Jumlah', 'Total Pendapatan']
    service_widths = [5, 30, 15, 12, 20]
    
    for col, (header, width) in enumerate(zip(service_headers, service_widths), start=1):
        cell = ws_service.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_service.column_dimensions[get_column_letter(col)].width = width
    
    service_stats = orders.values('service__name', 'service__type').annotate(
        count=Count('id'),
        total=Sum('price_total')
    ).order_by('-total')
    
    for idx, stat in enumerate(service_stats, start=1):
        row = idx + 3
        ws_service.cell(row=row, column=1, value=idx)
        ws_service.cell(row=row, column=2, value=stat['service__name'] or '-')
        ws_service.cell(row=row, column=3, value='Kiloan' if stat['service__type'] == 'per_kilo' else 'Satuan')
        ws_service.cell(row=row, column=4, value=stat['count'])
        ws_service.cell(row=row, column=5, value=float(stat['total'])).number_format = '#,##0'
    
    footer_service_row = len(service_stats) + 4
    if len(service_stats) > 0:
        ws_service.merge_cells(f'A{footer_service_row}:D{footer_service_row}')
        ws_service.cell(row=footer_service_row, column=1, value="TOTAL").font = Font(bold=True)
        ws_service.cell(row=footer_service_row, column=5, value=round(total_keseluruhan, 2)).font = Font(bold=True)
        ws_service.cell(row=footer_service_row, column=5).number_format = '#,##0'
    
    # ============================================================
    # SHEET 4: REKAP PER PELANGGAN
    # ============================================================
    ws_customer = wb.create_sheet("Rekap per Pelanggan")
    
    try:
        ws_customer.page_setup.orientation = 'portrait'
        ws_customer.page_setup.paperSize = '9'
        ws_customer.page_setup.fitToPage = True
        ws_customer.page_setup.fitToWidth = 1
    except:
        pass
    
    ws_customer.merge_cells('A1:E1')
    ws_customer['A1'] = f"REKAP PER PELANGGAN - {bulan_names[target_month]} {target_year}"
    ws_customer['A1'].font = Font(size=12, bold=True, color='1E40AF')
    ws_customer['A1'].alignment = Alignment(horizontal='center')
    
    customer_headers = ['No', 'Nama Pelanggan', 'Telepon', 'Jumlah Pesanan', 'Total Transaksi']
    customer_widths = [5, 25, 15, 15, 20]
    
    for col, (header, width) in enumerate(zip(customer_headers, customer_widths), start=1):
        cell = ws_customer.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_customer.column_dimensions[get_column_letter(col)].width = width
    
    customer_stats = orders.values('customer__username', 'customer__phone').annotate(
        count=Count('id'),
        total=Sum('price_total')
    ).order_by('-total')
    
    for idx, stat in enumerate(customer_stats, start=1):
        row = idx + 3
        ws_customer.cell(row=row, column=1, value=idx)
        ws_customer.cell(row=row, column=2, value=stat['customer__username'] or '-')
        ws_customer.cell(row=row, column=3, value=stat['customer__phone'] or '-')
        ws_customer.cell(row=row, column=4, value=stat['count'])
        ws_customer.cell(row=row, column=5, value=float(stat['total'])).number_format = '#,##0'
    
    # ============================================================
    # SHEET 5: STATISTIK
    # ============================================================
    ws_stats = wb.create_sheet("Statistik")
    
    try:
        ws_stats.page_setup.orientation = 'portrait'
        ws_stats.page_setup.paperSize = '9'
        ws_stats.page_setup.fitToPage = True
        ws_stats.page_setup.fitToWidth = 1
    except:
        pass
    
    ws_stats.merge_cells('A1:B1')
    ws_stats['A1'] = f"📈 STATISTIK TRANSAKSI - {bulan_names[target_month]} {target_year}"
    ws_stats['A1'].font = Font(size=12, bold=True, color='1E40AF')
    ws_stats['A1'].alignment = Alignment(horizontal='center')
    
    stats_data = [
        ('📦 Total Pesanan Lunas', f"{total_orders}"),
        ('⚖️ Total Berat Laundry', f"{total_berat:,.1f} kg".replace(',', '.')),
        ('💰 Total Pendapatan', f"Rp {total_keseluruhan:,.0f}".replace(',', '.')),
        ('🚚 Total Ongkir', f"Rp {total_ongkir:,.0f}".replace(',', '.')),
        ('🎉 Total Diskon', f"Rp {total_diskon:,.0f}".replace(',', '.')),
        ('📊 Rata-rata per Pesanan', f"Rp {(total_keseluruhan/total_orders if total_orders > 0 else 0):,.0f}".replace(',', '.')),
        ('⚖️ Rata-rata Berat per Pesanan', f"{(total_berat/total_orders if total_orders > 0 else 0):,.1f} kg".replace(',', '.')),
        ('🏷️ Jumlah Layanan', f"{len(service_stats)}"),
        ('👤 Jumlah Pelanggan', f"{customer_stats.count()}"),
        ('📸 Total Bukti Pembayaran', f"{sum(1 for o in orders if o.payment_proof)}"),
    ]
    
    for idx, (label, value) in enumerate(stats_data, start=3):
        ws_stats.cell(row=idx, column=1, value=label).font = Font(bold=True, size=11)
        ws_stats.cell(row=idx, column=2, value=value).font = Font(size=11, color='1E40AF')
        ws_stats.cell(row=idx, column=2).alignment = Alignment(horizontal='right')
        if idx % 2 == 0:
            ws_stats.cell(row=idx, column=1).fill = blue_fill
            ws_stats.cell(row=idx, column=2).fill = blue_fill
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 25
    
    # ============================================================
    # RESPONSE
    # ============================================================
    filename = f"Laporan_Transaksi_Lunas_{bulan_names[target_month]}_{target_year}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response